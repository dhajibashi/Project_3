# GLobal variables for the URLs
URL_FF3_MONTHLY = 'https://mba.tuck.dartmouth.edu/pages/faculty/ken.french/ftp/F-F_Research_Data_Factors_CSV.zip'
URL_KF_25_MONTHLY = 'https://mba.tuck.dartmouth.edu/pages/faculty/ken.french/ftp/25_Portfolios_5x5_CSV.zip'

# Global variables for the local file names
FF3_MONTHLY_FILE_NAME = 'F-F_Research_Data_Factors.csv'
KF_25_MONTHLY_FILE_NAME = '25_Portfolios_5x5.csv'

# Descriptive stats mode
USE_RAW_FOR_DISCRIPTIVES = False   # True to show Mean/StdDev on raw returns

# Model parameters
FF3_L = 3
CAPM_L = 1

# Imports
import os
import re
import zipfile
import requests
import warnings
from io import BytesIO, StringIO
from datetime import datetime

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import statsmodels.api as sm
from numpy.linalg import inv
from scipy.stats import f as f_dist
from scipy.stats import t as t_dist
from IPython.display import display

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

warnings.filterwarnings("ignore")


def prompt_date_range(dates):
    """
    Prompt user for a start and end date, normalize to YYYY-MM endpoints, and validate against the given date index.
    Keeps looping until valid input is given for both.
    """
    # Get min and max dates
    min_date = dates.min()
    max_date = dates.max()

    def normalize_date(s):
        """Try to convert user input to 'YYYY-MM' or Timestamp."""
        # Strip whitespace
        s = s.strip()
        # Patterns to match various formats
        patterns = [
            ("%Y-%m", r"^\d{4}-\d{2}$"),
            ("%m/%Y", r"^\d{2}/\d{4}$"),
            ("%Y%m",  r"^\d{6}$"),
            ("%b-%Y", r"^[A-Za-z]{3}-\d{4}$"),
            ("%B-%Y", r"^[A-Za-z]+-\d{4}$"),
        ]
        # Check for exact matches first
        for fmt, pat in patterns:
            # If matches pattern, try to parse to datetime
            if re.match(pat, s):
                try:
                    return pd.to_datetime(datetime.strptime(s, fmt)), None
                # If parsing fails, return an error message
                except Exception:
                    return None, f"Could not parse '{s}'"
        # If no pattern matched, try generic parsing
        try:
            return pd.to_datetime(s), None
        except Exception:
            return None, f"Could not parse '{s}'"

    while True:
        # Get user input for start date
        start_input = input(f"Enter start date on or after {min_date.strftime('%Y-%m')} (e.g., 2010-01): ").strip()
        # Try to parse start date
        start, err_start = normalize_date(start_input)
        # If error, print message and retry
        if err_start:
            print(err_start)
            continue
        
        # Get user input for end date
        end_input = input(f"Enter end date on or before {max_date.strftime('%Y-%m')} (e.g., 2023-12): ").strip()
        # Try to parse end date
        end, err_end = normalize_date(end_input)
        # If error, print message and retry
        if err_end:
            print(err_end)
            continue

        # Validate range
        # If the end date is before the start date, print message and retry
        if end < start:
            print("End date must be after or equal to start date.")
            continue

        # If either date is out of range, print message and retry
        if start < min_date or end > max_date:
            print(f"Date range must be within {min_date.strftime('%Y-%m')} and {max_date.strftime('%Y-%m')}.")
            continue

        # Normalize to month end
        start = pd.to_datetime(start) + pd.offsets.MonthEnd(0)
        end = pd.to_datetime(end) + pd.offsets.MonthEnd(0)

        # Show validated range and return
        print(f"Validated date range: {start.date()} to {end.date()}\n")
        return start, end
        

def prompt_model_choice():
    """
    Prompt user for model choice. Options: CAPM, FF3F, Both (case-insensitive).
    Returns:
        One of: 'CAPM', 'FF3F', 'Both'
    """
    # Valid options mapping
    valid = {"capm": "CAPM", "ff3f": "FF3F", "both": "Both"}
    
    while True:
        # Get user input
        inp = input("Choose model (CAPM, FF3F, Both): ").strip().lower()
        # Check if input is valid
        if inp in valid:
            print(f"Selected model: {valid[inp]}")
            return valid[inp]
        # If invalid, print message and retry
        print("Invalid option. Please type CAPM, FF3F, or Both.")


def download_and_save_kf_data(url: str, save_dir: str = "."):
    """
    Download Ken French dataset ZIP file, extract all CSV files, and save to the specified directory.

    Args:
        url (str): URL to the Ken French ZIP file.
        save_dir (str): Directory to save extracted CSVs (default: current folder).
    Returns:
        List of saved file paths.
    """
    # Make sure save_dir exists
    os.makedirs(save_dir, exist_ok=True)
    
    # Download the ZIP file
    resp = requests.get(url)
    resp.raise_for_status()

    # Extract CSV files from the ZIP
    with zipfile.ZipFile(BytesIO(resp.content)) as z:
        saved_files = []
        # Extract only CSV files
        for fname in z.namelist():
            if fname.endswith('.csv'):
                out_path = os.path.join(save_dir, os.path.basename(fname))
                # Extract and save
                with z.open(fname) as f_in, open(out_path, 'wb') as f_out:
                    f_out.write(f_in.read())
                saved_files.append(out_path)
    
    return saved_files


# ---------- small statistical helpers ----------
def monthly_sharpe(excess_series: pd.Series) -> float:
    # Sharpe ratio: mean / stddev of excess returns (monthly)
    s = excess_series.dropna()
    if s.shape[0] < 2:
        return np.nan
    mu = s.mean()
    sd = s.std(ddof=1)
    return np.nan if sd == 0 else mu / sd


def tstat_of_mean(series: pd.Series) -> float:
    # t-statistic for mean = mu / (sd/sqrt(T))
    s = series.dropna()
    T = s.shape[0]
    if T < 2:
        return np.nan
    mu = s.mean()
    sd = s.std(ddof=1)
    if sd == 0:
        return np.nan
    se = sd / np.sqrt(T)
    if se == 0:
        return np.nan
    return mu / se


def ols_with_const(y: pd.Series, X: pd.DataFrame):
    # OLS regression of y on X with constant term, handling missing data by dropping.
    Xc = sm.add_constant(X)
    return sm.OLS(y, Xc, missing="drop").fit()


def grs_test_with_alphas(alphas, factor_means, factor_cov, residual_cov, T, N, L):
    """
    Gibbons–Ross–Shanken (1989) F-stat for joint α = 0 across N assets, L factors.
    Returns (F, pval).
    """
    # Convert inputs to numpy arrays
    a = np.asarray(alphas, dtype=float).reshape(-1, 1)
    m = np.asarray(factor_means, dtype=float).reshape(-1, 1)
    # Calculate inverses of covariance matrices
    Sigma_f_inv = inv(np.asarray(factor_cov, dtype=float))
    Sigma_e_inv = inv(np.asarray(residual_cov, dtype=float))

    # Compute the GRS F-statistic
    term = float(m.T @ Sigma_f_inv @ m)
    numer = T / N * (T - N - L) / (T - L - 1)
    denom = 1.0 + term

    F = numer * float(a.T @ Sigma_e_inv @ a) / denom
    # Degrees of freedom for F-test
    df1, df2 = N, T - N - L
    if df2 <= 0:
        return np.nan, np.nan
    # p-value from F-distribution
    pval = 1.0 - f_dist.cdf(F, df1, df2)
    return F, pval


def _first_monthly_index_csv(lines):
    """
    Return the line index where monthly data start (first token YYYYMM), 
    assuming comma-separated rows.
    """
    for i, line in enumerate(lines):
        # Split by commas and strip whitespace
        toks = [t.strip() for t in line.strip().split(",")]
        # Check if first token is a 6-digit number (YYYYMM)
        # If so, return the index
        if len(toks) > 0 and toks[0].isdigit() and len(toks[0]) == 6:
            return i
    return None


def _dedupe(names):
    """
    Make column names unique by appending _1, _2, ... to duplicates.
    """

    # Use a dictionary to count occurrences
    out, seen = [], {}
    for n in names:
        # Normalize empty names to "COL"
        n = (n or "").strip()
        if n == "":
            n = "COL"
        # If seen before, append count; else initialize count
        if n in seen:
            seen[n] += 1
            out.append(f"{n}_{seen[n]}")
        else:
            seen[n] = 0
            out.append(n)
    return out


def _align_header_to_ncols(header, ncols):
    """
    Ensure header length == ncols; pad or truncate as needed.
    """
    # Convert to list if not already
    header = list(header)
    # Pad with COLx or truncate to match ncols
    if len(header) < ncols:
        pad = [f"COL{j}" for j in range(len(header)+1, ncols+1)]
        header = header + pad
    # Truncate if too long
    elif len(header) > ncols:
        header = header[:ncols]
    return header


# ---------- read Ken French 25-portfolios (monthly) from CSV (no Path) ----------
def read_kf_25_csv(filename: str) -> pd.DataFrame:
    # Open the file and read all lines
    with open(filename, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.readlines()

    # Find the line index where the row contains 'Average Equal Weighted Returns -- Monthly'
    filter_start = None
    for i, line in enumerate(lines):
        if 'Average Equal Weighted Returns -- Monthly' in line:
            filter_start = i + 1
            break

    # If not found, raise error
    if filter_start is None:
        raise ValueError("Could not find line containing 'Average Equal Weighted Returns -- Monthly'")

    # Now find first monthly data index after this line
    start_idx = _first_monthly_index_csv(lines[filter_start:])
    if start_idx is None:
        raise ValueError("Could not locate YYYYMM data start in 25_Portfolios_5x5.csv")

    # Adjust start_idx to be relative to full lines
    start_idx = start_idx + filter_start

    # Find end index for data block: first empty line or where first token is not YYYYMM format
    end_idx = len(lines)
    for j in range(start_idx, len(lines)):
        first_token = lines[j].strip().split(",")[0]
        if not (first_token.isdigit() and len(first_token) == 6):
            end_idx = j
            break

    # Header is previous line; parse by commas (NOT whitespace)
    header_line = lines[start_idx - 1]
    header_raw = [t.strip() for t in header_line.strip().split(",")]

    # Detect ncols from the FIRST data row (comma-split)
    first_data_tokens = [t.strip() for t in lines[start_idx].strip().split(",")]
    ncols = len(first_data_tokens)

    # Force first column to 'YYYYMM'
    if not header_raw:
        header_raw = ["YYYYMM"]
    else:
        header_raw[0] = "YYYYMM"

    # align and dedupe
    header = _align_header_to_ncols(header_raw, ncols)
    header = _dedupe(header)

    # read the data block using comma separator with slice lines
    text = "".join(lines[start_idx:end_idx])
    df = pd.read_csv(StringIO(text), sep=",", engine="python", header=None, names=header)

    # keep only proper YYYYMM rows
    df = df[df["YYYYMM"].astype(str).str.isdigit()].copy()
    df["YYYYMM"] = df["YYYYMM"].astype(int)
    df["date"] = pd.to_datetime(df["YYYYMM"].astype(str) + "01", format="%Y%m%d") + pd.offsets.MonthEnd(0)

    # set index and drop YYYYMM
    df = df.set_index("date").drop(columns=["YYYYMM"])

    # keep first 25 portfolio columns (leftmost 25 after YYYYMM)
    keep_cols = list(df.columns)[:25]
    df = df[keep_cols].apply(pd.to_numeric, errors="coerce") / 100.0
    return df


# ---------- read FF factors (monthly only) from CSV (no Path) ----------
def read_ff3_monthly(filename: str) -> pd.DataFrame:
    # Open the file and read all lines
    with open(filename, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.readlines()

    # find where Annual section starts (truncate monthly section there)
    stop_idx = None
    # Look for line containing "Annual" or "ANNUAL"
    for i, line in enumerate(lines):
        if "Annual" in line or "ANNUAL" in line:
            stop_idx = i
            break
    # Keep only lines up to stop_idx (if found)
    monthly_lines = lines if stop_idx is None else lines[:stop_idx]

    # Find first monthly data index
    start_idx = _first_monthly_index_csv(monthly_lines)
    # If not found, raise error
    if start_idx is None:
        raise ValueError("Could not locate YYYYMM monthly start in F-F_Research_Data_Factors.csv")

    # Header is previous line; parse by commas (NOT whitespace)
    header_line = monthly_lines[start_idx - 1]
    header_raw = [t.strip() for t in header_line.strip().split(",")]
    # Detect ncols from the FIRST data row (comma-split)
    first_data_tokens = [t.strip() for t in monthly_lines[start_idx].strip().split(",")]
    ncols = len(first_data_tokens)

    # first column
    if not header_raw:
        header_raw = ["YYYYMM"]
    else:
        header_raw[0] = "YYYYMM"

    # align and dedupe
    header = _align_header_to_ncols(header_raw, ncols)
    header = _dedupe(header)

    # read the data block using comma separator with slice lines
    text = "".join(monthly_lines[start_idx:])
    df = pd.read_csv(StringIO(text), sep=",", engine="python", header=None, names=header)

    # keep only monthly numeric rows
    df = df[df["YYYYMM"].astype(str).str.isdigit()].copy()
    df["YYYYMM"] = df["YYYYMM"].astype(int)
    df["date"] = pd.to_datetime(df["YYYYMM"].astype(str) + "01", format="%Y%m%d") + pd.offsets.MonthEnd(0)

    # rename factors to RMRF/SMB/HML/RF
    rename_map = {}
    for c in df.columns:
        cu = c.strip().upper()
        if "MKT" in cu: rename_map[c] = "RMRF"
        elif cu == "SMB": rename_map[c] = "SMB"
        elif cu == "HML": rename_map[c] = "HML"
        elif cu == "RF" : rename_map[c] = "RF"
    df = df.rename(columns=rename_map)

    # keep only relevant columns, set index
    keep = [c for c in ["date","RMRF","SMB","HML","RF"] if c in (["date"] + list(df.columns))]
    df = df[keep].set_index("date").sort_index()

    # numeric & % → decimal
    for c in [col for col in ["RMRF","SMB","HML","RF"] if col in df.columns]:
        df[c] = pd.to_numeric(df[c], errors="coerce") / 100.0
    return df


def normalize_port_names(cols):
    out = []
    for c in cols:
        # Replace SMALL→ME1, BIG→ME5, LoBM→BM1, HiBM→BM5
        s = str(c).strip()
        s = s.replace("SMALL", "ME1").replace("BIG", "ME5")
        s = s.replace("LoBM", "BM1").replace("HiBM", "BM5")
        s = re.sub(r"\s+", " ", s).upper()
        s = s.replace(" ", "_")
        out.append(s)
    return out


def normalize_and_rename_portfolios(data):
    # Identify factor and portfolio columns
    factor_cols = [c for c in ["RMRF", "SMB", "HML", "RF"] if c in data.columns]
    port_cols = [c for c in data.columns if c not in factor_cols]
    
    # Rename portfolio columns to standardized format
    new_port_cols = normalize_port_names(port_cols)
    data = data.rename(columns={old: new for old, new in zip(port_cols, new_port_cols)})
    
    # Re-identify factor_cols and port_cols with updated names
    factor_cols = [c for c in ["RMRF", "SMB", "HML", "RF"] if c in data.columns]
    port_cols = [c for c in data.columns if c not in factor_cols]

    return data, factor_cols, port_cols


def calculate_descriptive_statistics(sample, port_cols, excess, use_raw_for_descriptives):
    # Calculate descriptive statistics for portfolios
    # Use raw returns or excess returns based on flag
    base_for_stats = sample[port_cols] if use_raw_for_descriptives else excess
    rows = []
    for p in port_cols:
        rows.append({
            "Portfolio": p,
            "Mean (monthly)": base_for_stats[p].mean(),
            "StdDev (monthly)": base_for_stats[p].std(ddof=1),
            "Sharpe (monthly)": monthly_sharpe(excess[p]),
            "t(mean excess)": tstat_of_mean(excess[p]),
        })
    desc_table_monthly = pd.DataFrame(rows).set_index("Portfolio")

    # Annualize statistics
    desc_table_yearly = pd.DataFrame(index=desc_table_monthly.index)
    desc_table_yearly["Mean (yearly)"] = desc_table_monthly["Mean (monthly)"] * 12
    desc_table_yearly["StdDev (yearly)"] = desc_table_monthly["StdDev (monthly)"] * np.sqrt(12)
    desc_table_yearly["Sharpe (yearly)"] = desc_table_monthly["Sharpe (monthly)"] * np.sqrt(12)
    # t-stat remains the same
    desc_table_yearly["t(mean excess)"] = desc_table_monthly["t(mean excess)"].copy()
    return desc_table_monthly, desc_table_yearly


def cov_div_T_minus_L(X, L):
    T = X.shape[0]
    # Center data
    X_centered = X - X.mean(axis=0)
    # Compute covariance matrix with divisor T-L
    cov_mat = (X_centered.T @ X_centered) / (T - L)
    return cov_mat


def run_capm_regression(sample, port_cols, excess, CAPM_L):
    capm_rows, alphas_capm, resids_capm = [], [], []
    # CAPM uses only RMRF
    X_capm = sample[["RMRF"]]
    # Get dimensions
    T_capm = sample.shape[0]
    # Number of assets
    N_capm = len(port_cols)
    # Number of factors
    L_capm = CAPM_L

    # Run CAPM regression for each portfolio
    for p in port_cols:
        y = excess[p]
        # OLS regression with constant
        m = ols_with_const(y, X_capm)

        a    = m.params.get("const", np.nan)
        b    = m.params.get("RMRF", np.nan)
        se_a = m.bse.get("const", np.nan)
        se_b = m.bse.get("RMRF", np.nan)
        t_a  = m.tvalues.get("const", np.nan)
        t_b  = m.tvalues.get("RMRF", np.nan)
        r2   = m.rsquared

        capm_rows.append({
            "Portfolio": p,
            "Alpha": a, "SE(Alpha)": se_a, "t(Alpha)": t_a,
            "Beta_MKT": b, "SE(Beta_MKT)": se_b, "t(Beta_MKT)": t_b,
            "R2": r2
        })
        alphas_capm.append(a)
        resids_capm.append(m.resid)

    capm_table = pd.DataFrame(capm_rows).set_index("Portfolio")
    avg_r2_capm = capm_table["R2"].mean()

    # residual covariance (N×N)
    resid_mat_capm = pd.DataFrame({p: resids_capm[i] for i, p in enumerate(port_cols)}).dropna()
    # Covariance matrix of residuals with divisor T-L
    Sigma_e_capm   = cov_div_T_minus_L(resid_mat_capm, L_capm)

    # factor means/cov
    f_means_capm = [sample["RMRF"].mean()]
    # Covariance matrix of factor with divisor T-L
    f_cov_capm   = cov_div_T_minus_L(sample[["RMRF"]], L_capm)

    # GRS test
    GRS_capm, p_capm = grs_test_with_alphas(alphas_capm, f_means_capm, f_cov_capm, Sigma_e_capm, T_capm, N_capm, L_capm)

    capm_summary = pd.DataFrame({
        "Average R2": [avg_r2_capm],
        "GRS (CAPM)": [GRS_capm],
        "p-value": [p_capm],
        "T": [T_capm],
        "N": [N_capm]
    })
    capm_summary.reset_index(drop=True, inplace=True)
    # Name the row for clarity
    capm_summary.index = ["CAPM"]

    return capm_table, capm_summary, f_cov_capm, Sigma_e_capm


def run_ff3f_regression(sample, port_cols, excess, FF3_L):
    ff3_rows, alphas_ff3, resids_ff3 = [], [], []
    # FF3F uses RMRF, SMB, HML
    X_ff3 = sample[["RMRF", "SMB", "HML"]]

    T_ff3 = sample.shape[0]
    N_ff3 = len(port_cols)
    L_ff3 = FF3_L

    # Run FF3F regression for each portfolio
    for p in port_cols:
        y = excess[p]
        m = ols_with_const(y, X_ff3)

        a     = m.params.get("const", np.nan)
        b_mkt = m.params.get("RMRF", np.nan)
        b_smb = m.params.get("SMB",  np.nan)
        b_hml = m.params.get("HML",  np.nan)

        se_a  = m.bse.get("const", np.nan)
        se_m  = m.bse.get("RMRF",  np.nan)
        se_s  = m.bse.get("SMB",   np.nan)
        se_h  = m.bse.get("HML",   np.nan)

        t_a   = m.tvalues.get("const", np.nan)
        t_m   = m.tvalues.get("RMRF",  np.nan)
        t_s   = m.tvalues.get("SMB",   np.nan)
        t_h   = m.tvalues.get("HML",   np.nan)

        r2    = m.rsquared

        ff3_rows.append({
            "Portfolio": p,
            "Alpha": a, "SE(Alpha)": se_a, "t(Alpha)": t_a,
            "Beta_MKT": b_mkt, "SE(Beta_MKT)": se_m, "t(Beta_MKT)": t_m,
            "Beta_SMB": b_smb, "SE(Beta_SMB)": se_s, "t(Beta_SMB)": t_s,
            "Beta_HML": b_hml, "SE(Beta_HML)": se_h, "t(Beta_HML)": t_h,
            "R2": r2
        })
        alphas_ff3.append(a)
        resids_ff3.append(m.resid)

    ff3_table = pd.DataFrame(ff3_rows).set_index("Portfolio")
    avg_r2_ff3 = ff3_table["R2"].mean()

    # residual covariance
    resid_mat_ff3 = pd.DataFrame({p: resids_ff3[i] for i, p in enumerate(port_cols)}).dropna()
    Sigma_e_ff3   = cov_div_T_minus_L(resid_mat_ff3, L_ff3)

    # factor moments
    f_means_ff3 = sample[["RMRF", "SMB", "HML"]].mean().values
    f_cov_ff3   = cov_div_T_minus_L(sample[["RMRF", "SMB", "HML"]], L_ff3)

    # GRS test
    GRS_ff3, p_ff3 = grs_test_with_alphas(alphas_ff3, f_means_ff3, f_cov_ff3, Sigma_e_ff3, T_ff3, N_ff3, L_ff3)

    ff3_summary = pd.DataFrame({
        "Average R2": [avg_r2_ff3],
        "GRS (FF3F)": [GRS_ff3],
        "p-value": [p_ff3],
        "T": [T_ff3],
        "N": [N_ff3]
    })
    ff3_summary.reset_index(drop=True, inplace=True)
    # Name the row for clarity
    ff3_summary.index = ["FF3F"]

    return ff3_table, ff3_summary, f_cov_ff3, Sigma_e_ff3


def fmb_procedure(excess_returns: pd.DataFrame, betas: pd.DataFrame) -> pd.DataFrame:
    """
    Fama-MacBeth procedure for estimating risk premia.
    1. For each time period t, run cross-sectional regression of excess returns on betas to estimate risk premia.
    2. Obtain time series of risk premia estimates and compute their means, standard errors, and t-stats.
    3. Return a summary DataFrame with the results.

    Parameters:
    - excess_returns: pd.DataFrame, shape (T, N), excess returns for N assets over T periods
    - betas: pd.DataFrame, shape (N, K), factor loadings (betas)

    Returns:
    - summary_df: pd.DataFrame with columns [Mean, StdErr, t-Stat] indexed by factor name (+ Alpha).
    """
    # Get dimensions T, N, K
    T, N = excess_returns.shape
    K = betas.shape[1]

    risk_premia_estimates = []

    # For each time period, run cross-sectional regression
    for t in range(T):
        y = excess_returns.iloc[t].values
        X = betas.values

        # Add constant (intercept) term
        m = ols_with_const(y, pd.DataFrame(X))

        # Extract estimated risk premia (lambdas)
        lambda_t = m.params

        # Append parameters including intercept
        risk_premia_estimates.append(lambda_t.values if lambda_t is not None else [np.nan] * (K + 1))

    risk_premia_estimates = np.array(risk_premia_estimates)  # shape (T, K+1)

    # Compute time-series mean, std error, and t-stats
    mean_lambda = np.nanmean(risk_premia_estimates, axis=0)
    std_err_lambda = np.nanstd(risk_premia_estimates, axis=0, ddof=1) / \
        np.sqrt(np.sum(~np.isnan(risk_premia_estimates), axis=0))
    t_stats = mean_lambda / std_err_lambda
    # Two-tailed p-values
    pval = 2 * (1 - t_dist.cdf(np.abs(t_stats), df=T-1))

    # Add "Alpha" as intercept name + factor names
    factor_names = ["Alpha"] + list(betas.columns)

    summary_df = pd.DataFrame({
        "Mean": mean_lambda,
        "StdErr": std_err_lambda,
        "t-Stat": t_stats,
        "p-value": pval
    }, index=factor_names)

    return summary_df


def shanken_correction(summary_df: pd.DataFrame, factor: pd.DataFrame) -> pd.DataFrame:
    '''
    Apply the Shanken correction to the standard errors and t-stats of risk premia estimates,
    including the alpha (intercept).

    Parameters:
    - summary_df: DataFrame with columns ['Mean', 'StdErr', 't-Stat'] indexed by factor names,
                  first index is assumed 'Alpha'.
    - factor: DataFrame of factor returns (T x K).

    Returns:
    - corrected_summary_df: DataFrame with updated 'StdErr' and 't-Stat'.
    '''
    # Get dimensions L, T
    L = factor.shape[1]
    T = factor.shape[0]

    # Compute factor covariance matrix with divisor T-L
    factorcovmat = cov_div_T_minus_L(factor, L)

    # Extract means of all lambdas including alpha
    # Assume index order: alpha first, then factors
    lambdas_all = summary_df["Mean"].values.reshape(-1, 1)
    
    # Separate alpha and factor lambdas
    alpha_lambda = lambdas_all[0:1]
    factor_lambdas = lambdas_all[1:]

    # Compute correction term for factors only
    inv_factorcov = np.linalg.inv(factorcovmat)
    factor_correction = 1 + (factor_lambdas.T @ inv_factorcov @ factor_lambdas).item()

    # Apply the same correction factor to alpha and factor std errors
    corrected_std_err = summary_df["StdErr"] * np.sqrt(factor_correction)
    corrected_t_stat = summary_df["Mean"] / corrected_std_err

    corrected_summary_df = summary_df.copy()
    corrected_summary_df["StdErr"] = corrected_std_err
    corrected_summary_df["t-Stat"] = corrected_t_stat
    # Recompute two-tailed p-values
    corrected_summary_df["p-value"] = 2 * (1 - t_dist.cdf(np.abs(corrected_t_stat), df=T-1))

    return corrected_summary_df


def aggregate_results(fmb_summary: pd.DataFrame, fmb_summary_shanken: pd.DataFrame) -> pd.DataFrame:
    """
    Combine Fama-MacBeth summary and Shanken-corrected summary into a single DataFrame for display.
    """
    # Create a copy of the original summary and rename columns to indicate uncorrected and Shanken-corrected
    agg = fmb_summary.copy()
    agg.columns = [f"{col} (uncorrected)" for col in agg.columns]
    shanken_cols = {col: f"{col} (Shanken)" for col in fmb_summary_shanken.columns}
    agg = agg.join(fmb_summary_shanken.rename(columns=shanken_cols))
    return agg


def plot_capm_alpha_bars(capm_table):
    '''
    Plot horizontal bar chart of CAPM alphas by portfolio.
    '''
    fig, ax = plt.subplots()
    capm_table["Alpha"].sort_values().plot(kind="barh", ax=ax)
    ax.set_title("CAPM Alphas by Portfolio (monthly)")
    ax.set_xlabel("Alpha")
    fig.tight_layout()
    return fig


def plot_ff3f_alpha_bars(ff3_table):
    '''
    Plot horizontal bar chart of FF3F alphas by portfolio.
    '''
    fig, ax = plt.subplots()
    ff3_table["Alpha"].sort_values().plot(kind="barh", ax=ax)
    ax.set_title("FF3F Alphas by Portfolio (monthly)")
    ax.set_xlabel("Alpha")
    fig.tight_layout()
    return fig


def plot_r2_scatter_capm_vs_ff3f(capm_table, ff3_table):
    '''
    Scatter plot comparing R² from CAPM vs FF3F for each portfolio.
    '''
    fig, ax = plt.subplots()
    x = capm_table["R2"]
    y = ff3_table["R2"]
    ax.scatter(x, y)
    minv = float(min(x.min(), y.min()))
    maxv = float(max(x.max(), y.max()))
    # Create a 45-degree reference line
    ax.plot([minv, maxv], [minv, maxv])
    ax.set_title("R²: CAPM vs FF3F (per portfolio)")
    ax.set_xlabel("R² CAPM")
    ax.set_ylabel("R² FF3F")
    fig.tight_layout()
    return fig


def plot_fmb_capm_lambdas(fmb_capm_summary):
    '''
    Plot FMB CAPM lambdas (Alpha and Beta_MKT) with error bars.
    '''
    fig, ax = plt.subplots()
    s = fmb_capm_summary.loc[["Alpha", "Beta_MKT"]]
    x_pos = [0.5, 1.5]
    ax.errorbar(x_pos, s["Mean"], yerr=s["StdErr"], fmt='o')
    ax.set_xticks(x_pos)
    ax.set_xticklabels(s.index, rotation=0)
    ax.set_xlim(0, 2)
    ax.set_title("FMB CAPM — Lambda Means with StdErr")
    fig.tight_layout()
    return fig


def plot_fmb_ff3f_lambdas(fmb_ff3_summary):
    '''
    Plot FMB FF3F lambdas (Alpha, Beta_MKT, Beta_SMB, Beta_HML) with error bars.
    '''
    fig, ax = plt.subplots()
    s = fmb_ff3_summary.loc[["Alpha", "Beta_MKT", "Beta_SMB", "Beta_HML"]]
    ax.errorbar(range(s.shape[0]), s["Mean"], yerr=s["StdErr"], fmt='o')
    ax.set_xticks(range(s.shape[0]))
    ax.set_xticklabels(s.index, rotation=0)
    ax.set_title("FMB FF3F — Lambda Means with StdErr")
    fig.tight_layout()
    return fig


def run_model(model, sample, port_cols, excess, CAPM_L, FF3_L, use_raw_for_descriptives):
    # Calculate descriptive statistics of both monthly and yearly
    desc_table_monthly, desc_table_yearly = calculate_descriptive_statistics(sample, port_cols, excess, use_raw_for_descriptives)
    print("Descriptive Statistics of Portfolios (Yearly):")
    print(desc_table_yearly, '\n')
    print("Descriptive Statistics of Portfolios (Monthly):")
    print(desc_table_monthly, '\n')

    # Initialize all outputs to None
    capm_table = capm_summary = f_cov_capm = Sigma_e_capm = fmb_capm_summary = fmb_capm_summary_shanken = agg_capm = None
    ff3_table = ff3_summary = f_cov_ff3 = Sigma_e_ff3 = fmb_ff3_summary = fmb_ff3_summary_shanken = agg_ff3 = None

    if model in ("CAPM", "Both"):
        # Run CAPM regression
        capm_table, capm_summary, f_cov_capm, Sigma_e_capm = run_capm_regression(sample, port_cols, excess, CAPM_L)
        print("CAPM Regression Results:")
        print(capm_table, '\n')
        print("CAPM GRS Test Summary:")
        print(capm_summary, '\n')

        # Fama-MacBeth procedure for CAPM
        fmb_capm_summary = fmb_procedure(excess, capm_table[["Beta_MKT"]])

        # Shanken correction for CAPM
        fmb_capm_summary_shanken = shanken_correction(fmb_capm_summary, sample[["RMRF"]])

        # Aggregate results and display
        agg_capm = aggregate_results(fmb_capm_summary, fmb_capm_summary_shanken)
        print("FMB CAPM Summary (with Shanken correction):")
        print(agg_capm, '\n')

    if model in ("FF3F", "Both"):
        # Run FF3F regression
        ff3_table, ff3_summary, f_cov_ff3, Sigma_e_ff3 = run_ff3f_regression(sample, port_cols, excess, FF3_L)
        print("FF3F Regression Results:")
        print(ff3_table, '\n')
        print("FF3F GRS Test Summary:")
        print(ff3_summary, '\n')

        # Fama-MacBeth procedure for FF3F
        fmb_ff3_summary = fmb_procedure(excess, ff3_table[["Beta_MKT", "Beta_SMB", "Beta_HML"]])
        
        # Shanken correction for FF3F
        fmb_ff3_summary_shanken = shanken_correction(fmb_ff3_summary, sample[["RMRF","SMB","HML"]])
        
        # Aggregate results and display
        agg_ff3 = aggregate_results(fmb_ff3_summary, fmb_ff3_summary_shanken)
        print("FMB FF3F Summary (with Shanken correction):")
        print(agg_ff3, '\n')

    # Generate charts based on selected model
    chart_outputs = {}
    if model in ("CAPM", "Both"):
        chart_outputs["plot_capm_alpha_bars"] = plot_capm_alpha_bars(capm_table)

    if model in ("FF3F", "Both"):
        chart_outputs["plot_ff3f_alpha_bars"] = plot_ff3f_alpha_bars(ff3_table)

    if model == "Both":
        chart_outputs["plot_r2_scatter_capm_vs_ff3f"] = plot_r2_scatter_capm_vs_ff3f(capm_table, ff3_table)
        chart_outputs["plot_fmb_capm_lambdas"] = plot_fmb_capm_lambdas(fmb_capm_summary)
        chart_outputs["plot_fmb_ff3f_lambdas"] = plot_fmb_ff3f_lambdas(fmb_ff3_summary)

    # Return all results in a dictionary
    return {
        "desc_table_monthly": desc_table_monthly,
        "desc_table_yearly": desc_table_yearly,
        "capm_table": capm_table,
        "capm_summary": capm_summary,
        "f_cov_capm": f_cov_capm,
        "Sigma_e_capm": Sigma_e_capm,
        "agg_capm": agg_capm,
        "ff3_table": ff3_table,
        "ff3_summary": ff3_summary,
        "f_cov_ff3": f_cov_ff3,
        "Sigma_e_ff3": Sigma_e_ff3,
        "agg_ff3": agg_ff3,
        "charts": chart_outputs
    }


def file_name_export(model, start, end):
    '''
    Generate a filename for the Excel export based on model and date range.
    '''
    # Format dates as 'YYYYMMDD'
    start_str = start.strftime('%Y%m')
    end_str = end.strftime('%Y%m')
    
    if model == "Both":
        model_part = "CAPM_FF3F"
    elif model == "CAPM":
        model_part = "CAPM"
    elif model == "FF3F":
        model_part = "FF3F"
    else:
        raise ValueError("Invalid model")
    
    # Return formatted filename
    return f"analysis_report_{model_part}_{start_str}_{end_str}.xlsx"


def export_to_excel(filename, input_info, results, charts=None):
    """
    Export all analysis results to a formatted Excel workbook.

    Parameters:
    - filename: output Excel file path
    - input_info: dict with input specs (date range, model, etc.)
    - results: dict with DataFrames of analysis results, e.g. 
      'desc_table_monthly', 'capm_table', 'capm_summary', 'f_cov_capm', 'Sigma_e_capm', 
      'fmb_capm_summary', 'fmb_capm_summary_shanken', 'agg_capm',
      'ff3_table', 'ff3_summary', 'f_cov_ff3', 'Sigma_e_ff3',
      'fmb_ff3_summary', 'fmb_ff3_summary_shanken', 'agg_ff3'
    - charts: optional dict of matplotlib figures keyed by sheet/figure name

    Produces an Excel workbook with main summary tab, exhibits, covariances, GRS tests,
    Fama-MacBeth results and diagnostic charts if present.
    """
    # Adjust model name for clarity
    input_info["model"] = "CAPM/FF3F" if model == "Both" else model

    # Helper function to write DataFrame to worksheet with title
    def write_df(ws, df, title=None, start_row=1):
        # If title is provided, write it in bold at the start_row
        if title:
            ws.cell(row=start_row, column=1, value=title)
            ws.cell(row=start_row, column=1).font = Font(bold=True, size=14)
            start_row += 2

        # Write DataFrame starting from start_row
        for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == start_row:
                    cell.font = Font(bold=True)
                # Apply center alignment
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Create workbook and main summary sheet
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Main Summary"

    # Make cell widths wider for readability
    for col in range(1, 20):
        ws_summary.column_dimensions[get_column_letter(col)].width = 20
    
    # Input information
    ws_summary.cell(row=1, column=1, value="Input Information and Model Specifications")
    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
    row_cursor = 3
    # Write each key-value pair from input_info
    for k, v in input_info.items():
        ws_summary.cell(row=row_cursor, column=1, value=k)
        ws_summary.cell(row=row_cursor, column=2, value=str(v))
        row_cursor += 1
    row_cursor += 1
    
    # Descriptive statistics
    if results.get('desc_table_yearly') is not None:
        write_df(ws_summary, results['desc_table_yearly'], title="Portfolio Descriptive Statistics (Yearly)", start_row=row_cursor)
        row_cursor += results['desc_table_yearly'].shape[0] + 5

    # CAPM tables and summaries
    if results.get('capm_table') is not None:
        write_df(ws_summary, results['capm_summary'], title="CAPM Summary Statistics", start_row=row_cursor)
        row_cursor += results['capm_summary'].shape[0] + 5
    if results.get('agg_capm') is not None:
        write_df(ws_summary, results['agg_capm'], title="FMB CAPM Risk Premia", start_row=row_cursor)
        row_cursor += results['agg_capm'].shape[0] + 5

    # FF3F tables and summaries
    if results.get('ff3_table') is not None:
        write_df(ws_summary, results['ff3_summary'], title="FF3F Summary Statistics", start_row=row_cursor)
        row_cursor += results['ff3_summary'].shape[0] + 5
    if results.get('agg_ff3') is not None:
        write_df(ws_summary, results['agg_ff3'], title="FMB FF3F Risk Premia", start_row=row_cursor)
        row_cursor += results['agg_ff3'].shape[0] + 5

    # Portfolio summary tab
    ws_portfolio = wb.create_sheet("Portfolio Summary")
    for col in range(1, 20):
        ws_portfolio.column_dimensions[get_column_letter(col)].width = 15
    row_cursor = 1
    if results.get('desc_table_yearly') is not None:
        # Write yearly descriptive statistics
        write_df(ws_portfolio, results['desc_table_yearly'], title="Portfolio Descriptive Statistics (Yearly)", start_row=row_cursor)
        row_cursor += results['desc_table_yearly'].shape[0] + 5
    if results.get('desc_table_monthly') is not None:
        # Write monthly descriptive statistics
        write_df(ws_portfolio, results['desc_table_monthly'], title="Portfolio Descriptive Statistics (Monthly)", start_row=row_cursor)
        row_cursor += results['desc_table_monthly'].shape[0] + 5

    # Regression Outputs tab
    ws_regression = wb.create_sheet("Regression Outputs")
    # Make cell widths wider for readability
    for col in range(1, 20):
        ws_regression.column_dimensions[get_column_letter(col)].width = 15
    row_cursor = 1
    if results.get('capm_table') is not None:
        # Write CAPM regression results
        write_df(ws_regression, results['capm_table'], title="CAPM Regression Results", start_row=row_cursor)
        row_cursor += results['capm_table'].shape[0] + 5
    if results.get('ff3_table') is not None:
        # Write FF3F regression results
        write_df(ws_regression, results['ff3_table'], title="FF3F Regression Results", start_row=row_cursor)
        row_cursor += results['ff3_table'].shape[0] + 5

    # Residual & Factor Covariance tab
    ws_cov = wb.create_sheet("Covariance Matrices")
    for col in range(1, 30):
        ws_cov.column_dimensions[get_column_letter(col)].width = 15
    row_cursor = 1
    if results.get('Sigma_e_capm') is not None:
        # Write CAPM residual covariance
        write_df(ws_cov, pd.DataFrame(results['Sigma_e_capm']), title="CAPM Residual Covariance", start_row=row_cursor)
        row_cursor += results['Sigma_e_capm'].shape[0] + 5
    if results.get('f_cov_capm') is not None:
        # Write CAPM factor covariance
        write_df(ws_cov, pd.DataFrame(results['f_cov_capm']), title="CAPM Factor Covariance", start_row=row_cursor)
        row_cursor += results['f_cov_capm'].shape[0] + 5
    if results.get('Sigma_e_ff3') is not None:
        # Write FF3F residual covariance
        write_df(ws_cov, pd.DataFrame(results['Sigma_e_ff3']), title="FF3F Residual Covariance", start_row=row_cursor)
        row_cursor += results['Sigma_e_ff3'].shape[0] + 5
    if results.get('f_cov_ff3') is not None:
        # Write FF3F factor covariance
        write_df(ws_cov, pd.DataFrame(results['f_cov_ff3']), title="FF3F Factor Covariance", start_row=row_cursor)
        row_cursor += results['f_cov_ff3'].shape[0] + 5

    # GRS test components tab
    ws_grs = wb.create_sheet("GRS Test Outputs")
    # Make cell widths wider for readability
    for col in range(1, 20):
        ws_grs.column_dimensions[get_column_letter(col)].width = 15
    row_cursor = 1
    if results.get('capm_summary') is not None: 
        # Write CAPM GRS test summary
        write_df(ws_grs, results['capm_summary'], title="CAPM GRS Test Summary", start_row=row_cursor)
        row_cursor += results['capm_summary'].shape[0] + 5
    if results.get('ff3_summary') is not None:
        # Write FF3F GRS test summary
        write_df(ws_grs, results['ff3_summary'], title="FF3F GRS Test Summary", start_row=row_cursor)
        row_cursor += results['ff3_summary'].shape[0] + 5

    # Fama–MacBeth tabs
    ws_fmb = wb.create_sheet("FMB Outputs")
    # Make cell widths wider for readability
    for col in range(1, 20):
        ws_fmb.column_dimensions[get_column_letter(col)].width = 20
    row_cursor = 1
    if results.get('agg_capm') is not None:
        # Write aggregated FMB CAPM results with Shanken correction
        write_df(ws_fmb, results['agg_capm'], title="FMB CAPM Summary", start_row=row_cursor)
        row_cursor += results['agg_capm'].shape[0] + 5
    if results.get('agg_ff3') is not None:
        # Write aggregated FMB FF3F results with Shanken correction
        write_df(ws_fmb, results['agg_ff3'], title="FMB FF3F Summary", start_row=row_cursor)
        row_cursor += results['agg_ff3'].shape[0] + 5

    # Charts tab
    if results.get('charts') is not None:
        ws_chart = wb.create_sheet("Charts")
        top_left_cell_row = 1
        img_files = []

        # Save each figure to a temporary file and keep track of paths
        for chart_name, fig in results['charts'].items():
            img_path = f"_temp_{chart_name}.png"
            fig.savefig(img_path, bbox_inches='tight')
            img_files.append((img_path, chart_name))
        
        # Insert images into the Charts worksheet
        for img_path, chart_name in img_files:
            img = XLImage(img_path)
            cell_position = f"A{top_left_cell_row}"
            ws_chart.add_image(img, cell_position)
            img_height = img.height if img.height else 200
            rows_to_advance = max(img_height // 20, 15)
            top_left_cell_row += rows_to_advance

    # Save the workbook
    wb.save(filename)
    print(f"Excel report saved to: {filename}")

    # Delete temporary image files
    for img_path, _ in img_files:
        os.remove(img_path)


if __name__ == "__main__":

    # Download data files and save locally
    download_and_save_kf_data(URL_FF3_MONTHLY)
    download_and_save_kf_data(URL_KF_25_MONTHLY)

    # Read data
    ret_5x5 = read_kf_25_csv(f'{KF_25_MONTHLY_FILE_NAME}')
    ff3     = read_ff3_monthly(f'{FF3_MONTHLY_FILE_NAME}')

    # align on common months
    data = ret_5x5.join(ff3, how="inner")
    print("Portfolios:", ret_5x5.shape, "| Factors:", ff3.shape)
    print("Merged data range:", data.index.min().date(), "→", data.index.max().date(), '\n')

    data, factor_cols, port_cols = normalize_and_rename_portfolios(data)

    # User inputs of date range and model choice
    start, end = prompt_date_range(data.index)
    model = prompt_model_choice()

    # subset and rebuild lists
    sample = data.loc[
        (data.index >= pd.to_datetime(start) + pd.offsets.MonthEnd(0)) &
        (data.index <= pd.to_datetime(end)   + pd.offsets.MonthEnd(0))
    ].copy()

    # Calculate excess returns
    excess = sample[port_cols].sub(sample["RF"], axis=0)

    # Print sample information
    print("Sample window:", sample.index.min().date(), "→", sample.index.max().date())
    print("# portfolios:", len(port_cols), "| # months:", sample.shape[0], '\n')

    # Run the selected model
    results = run_model(model, sample, port_cols, excess, CAPM_L, FF3_L, USE_RAW_FOR_DISCRIPTIVES)

    # Create filename
    file_name = file_name_export(model, start, end)
    # Metadata for export
    input_info = {
        "model": model,
        "start": start,
        "end": end
    }
    # Export all results to Excel
    export_to_excel(file_name, input_info, results)
    # ---------- end of main code ----------

